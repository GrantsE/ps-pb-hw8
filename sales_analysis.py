import pandas as pd
import collections
import datetime as dt
import openpyxl

excel_data = pd.read_excel('logs.xlsx', sheet_name = 'log', engine='openpyxl')

# Добавляем новый столбец, где будет отображаться месяц
excel_data['Месяц'] = excel_data['Дата посещения'].dt.month

wb = openpyxl.load_workbook(filename = 'report.xlsx')

sheet = wb['Лист1']


#Создаём функцию и передаем туда 3 параметра
#В 1-ый параметр вводим название столбца (Браузер или Купленные товары) из файла, во 2-ой - номер первой строки, а в 3-й - номер конечной строки
def fill_data(data_type, number_of_rows, end_of_rows):

    #Если 1-й параметр = Браузер, то выводим 7 популярных наименований из столбца "Браузер". В противном случае выводим 7 популярных товаров 
    if data_type == 'Браузер':

        #Выводим самые популярные наименования и количество их повторений
        names_counter = dict(collections.Counter(excel_data['Браузер']).most_common(7))

    else:
        #Разделяем данные из столбца Купленные товары на элементы. В качестве разделителя используем запятую
        excel_data['Купленные товары'] = excel_data['Купленные товары'].str.split(',')

        excel_dict = excel_data.to_dict(orient = 'records')

        man_woman = []
        months = []
        all_items = []

        #Обходим данные из файла logs. В переменную man_woman добавляем данные по полу, в months - по месяцу, в all_items - все купленные товары
        for data in excel_dict:
            for items in data['Купленные товары']:
                man_woman.append(data['Пол'])
                months.append(data['Месяц'])
                all_items.append(items)
        
        #Выводим самые популярные товары и количество их повторений
        names_counter = dict(collections.Counter(all_items).most_common(7))

        #Создаём отдельную таблицу и передаём элементы из man_woman, months, all_items
        df = pd.DataFrame({'Пол':man_woman, 'Месяц':months, 'Купленные товары':all_items})

        #Создаём список, где будут даны элементы м и ж, то есть мужской или женский пол
        gender = ['м', 'ж']

        #Обходим элементы gender и выводим наиболее и наименее востребованные товары среди мужчин и женщин
        for g in gender:
            gender_info = (df[df['Пол'] == g])
            if g == 'м':
                sheet.cell(row = 31, column = 2).value = gender_info['Купленные товары'].value_counts().index[0]
                sheet.cell(row = 33, column = 2).value = gender_info['Купленные товары'].value_counts(ascending = True).index[0]
            else:
                sheet.cell(row = 32, column = 2).value = gender_info['Купленные товары'].value_counts().index[0]
                sheet.cell(row = 34, column = 2).value = gender_info['Купленные товары'].value_counts(ascending = True).index[0]

    row = int(number_of_rows)

    #Создаём цикл и добавляем данные по популярным браузерам/куплленным товарам и их количеству в соответствии с номерами строк и столбцов
    while row <= int(end_of_rows):
        for k, v in names_counter.items():
            sheet.cell(row = row, column = 1).value = k
            sheet.cell(row = row, column = 2).value = v
            row += 1
    
    month = 1
    columns = 3
    total_sum = []

    #Создаём цикл, который будет действовать до тех пор, пока month <= 12 и columns <= 14
    while month <= 12 and columns <= 14:
        row_number = number_of_rows
        if data_type == 'Браузер':  
            month_data = (excel_data[excel_data['Месяц'] == month])
        else:
            month_data = (df[df['Месяц'] == month])
        #Обходим 7 самых популярных браузеров или товаров (в зависимости от передаваемого параметра) и передаём данные по количеству за каждый месяц
        for names in names_counter.keys(): 
            quantities = month_data[month_data[data_type] == names]     
            total_sum.append(int(quantities.count()[0]))     
            sheet.cell(row = row_number, column = columns).value = quantities.count()[0]    
            row_number += 1
        sheet.cell(row = row_number, column = columns).value = sum(total_sum)
        del total_sum[0: len(total_sum) + 1]
        columns += 1
        month += 1

    wb.save(filename = 'report.xlsx')
    wb.close()

    
fill_data('Браузер', 5, 11)
fill_data('Купленные товары', 19, 25)

print('Все данные добавлены в файл')